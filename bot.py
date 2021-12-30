import telebot
import config 
import random
import openpyxl
import datetime
from telebot import types

bot = telebot.TeleBot(config.Token)

test_for_chanell = False

channel_adress = config.Channel_adress

@bot.message_handler(commands = ['start'])
def start(message):
	user_name = '{0.first_name} {0.last_name}'.format(message.from_user)

	sti = open('st/sticker.webp','rb')
	bot.send_sticker(message.chat.id,sti)

	bot.send_message(channel_adress, '{0.first_name} зашел(ла) в "Пельменную"'.format(message.from_user))

	markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
	item1 = types.KeyboardButton('Кинем монетку')
	item4 = types.KeyboardButton('Камень, ножницы, бумага')
	item2 = types.KeyboardButton('Рандомное число')
	item3 = types.KeyboardButton('Рецепт пельменей')
	item5 = types.KeyboardButton('Сообщение в канал')

	markup.add(item1,item2,item3,item4,item5)

	bot.send_message(message.chat.id, 'Добро пожаловать в Пельменную, {0.first_name}!\nЧто будем делать?'.format(message.from_user),
		reply_markup=markup)

	journal = open('journal.txt','r')
	dictionary = eval(journal.read())
	journal.close()

	if user_name in dictionary.keys():
		line_in_tabel = dictionary[str(user_name)]
		table = openpyxl.load_workbook('tables/'+user_name+'.xlsx')
		sheet = table.active	
	else:
		journal = open('journal.txt','w')
		
		line_in_tabel = 2
		dictionary[str(user_name)] = 2
		journal.write(str(dictionary))
		
		table = openpyxl.Workbook()
		sheet = table.active
		sheet.title = str(message.from_user.id)
		
		sheet['A1'] = 'action'
		sheet['B1'] = 'date'

		journal.close()
	
	table.save('tables/{}.xlsx'.format(user_name))

@bot.message_handler(content_types=['text'])
def lalala(message):
	global test_for_chanell
	
	user_name = '{0.first_name} {0.last_name}'.format(message.from_user)
	journal = open('journal.txt', 'r')
	dictionary = eval(journal.read())
	journal.close()
	
	if message.chat.type == 'private':
		if message.text == 'Кинем монетку':

			markup = types.InlineKeyboardMarkup(row_width=2)
			item1 = types.InlineKeyboardButton('Орел', callback_data='o')
			item2 = types.InlineKeyboardButton('Решка', callback_data='r')
			markup.add(item1,item2)

			bot.send_message(message.chat.id, 'Загадай сторону' , reply_markup=markup)

			journal = open('journal.txt', 'w')
			line_in_tabel = dictionary[str(user_name)]

			table = openpyxl.load_workbook('tables/'+str(user_name)+'.xlsx')
			
			sheet = table.active

			bot.send_message(channel_adress, '{} кидает монетку'.format(user_name))

			sheet['A'+str(line_in_tabel)] = 'Кинул(а) монетку'
			sheet['B'+str(line_in_tabel)] = str(datetime.datetime.now())[0:19]

			table.save('tables/{}.xlsx'.format(user_name))

			line_in_tabel += 1
			dictionary[str(user_name)] = line_in_tabel
			
			journal.write(str(dictionary))
			journal.close()

		elif message.text == 'Рандомное число':

			journal = open('journal.txt', 'w')
			line_in_tabel = dictionary[str(user_name)]

			table = openpyxl.load_workbook('tables/'+str(user_name)+'.xlsx')
			
			sheet = table.active

			bot.send_message(channel_adress, '{} узнает рандомное число'.format(user_name))

			sheet['A'+str(line_in_tabel)] = 'Узнает рандомное число'
			sheet['B'+str(line_in_tabel)] = str(datetime.datetime.now())[0:19]

			table.save('tables/{}.xlsx'.format(user_name))

			line_in_tabel += 1
			dictionary[str(user_name)] = line_in_tabel
			
			journal.write(str(dictionary))
			journal.close()

			bot.send_message(message.chat.id, random.randint(1,100))
		elif message.text == 'Рецепт пельменей':
			
			sh1 = open('st/шаг 1.jpg','rb')
			sh2 = open('st/шаг 2.jpg','rb')
			sh3 = open('st/шаг 3.jpg','rb')
			sh4 = open('st/шаг 4.jpg','rb')
			sh5 = open('st/шаг 5.jpg','rb')
			sh6 = open('st/шаг 6.jpg','rb')

			journal = open('journal.txt', 'w')
			line_in_tabel = dictionary[str(user_name)]

			table = openpyxl.load_workbook('tables/'+str(user_name)+'.xlsx')
			
			sheet = table.active

			bot.send_message(channel_adress, '{} узнает рецепт пельменей'.format(user_name))

			sheet['A'+str(line_in_tabel)] = 'Узнает рецепт пельменей'
			sheet['B'+str(line_in_tabel)] = str(datetime.datetime.now())[0:19]

			table.save('tables/{}.xlsx'.format(user_name))

			line_in_tabel += 1
			dictionary[str(user_name)] = line_in_tabel
			
			journal.write(str(dictionary))
			journal.close()

			bot.send_message(message.chat.id,'ИНГРЕДИЕНТЫ\n*1 яйцо\n*1 стакан воды\n*1 ч. л. соли\n*600 г пшеничной муки\nДЛЯ НАЧИНКИ:\n*250 г говяжьего фарша\n*250 г свиного фарша\n*1 большая луковица\n*1 зубчик чеснока')
			
			bot.send_message(message.chat.id, 'ШАГ 1')
			bot.send_message(message.chat.id, 'Приготовить тесто для пельменей. Муку просеять горкой. Сделать сверху углубление, влить в него яйцо и 1 ст. л. воды, добавить щепотку соли.')
			bot.send_sticker(message.chat.id,sh1)

			bot.send_message(message.chat.id, 'ШАГ 2')
			bot.send_message(message.chat.id, 'Собирая муку с краев к середине, чтобы вода и яйцо не выливались из углубления, замешивать тесто, добавляя небольшими порциями оставшуюся воду. Месить тесто для пельменей до тех пор, пока оно не станет эластичным и однородным, примерно 10 мин. Накрыть влажным полотенцем и оставить на 30 мин. Вода для пельменного теста должна быть ледяная. Для этого ее заранее ставят в холодильник или морозильник.')
			bot.send_sticker(message.chat.id,sh2)

			bot.send_message(message.chat.id, 'ШАГ 3')
			bot.send_message(message.chat.id, 'Пока расстаивается тесто, приготовить начинку для пельменей. Лук и чеснок очистить и очень мелко порубить. Смешать говяжий и свиной фарш, добавить лук с чесноком, посолить, поперчить. Тщательно перемешать до однородности.')
			bot.send_sticker(message.chat.id,sh3)

			bot.send_message(message.chat.id, 'ШАГ 4')
			bot.send_message(message.chat.id, 'Готовое тесто для русских пельменей разделить на 4 части. 3 части накрыть влажным полотенцем и отложить. Оставшееся тесто скатать в жгут толщиной 2 см. Нарезать его на кусочки шириной 1,5 см. На присыпанной мукой поверхности раскатать каждый кусочек теста в тонкую лепешку.')
			bot.send_sticker(message.chat.id,sh4)

			bot.send_message(message.chat.id, 'ШАГ 5')
			bot.send_message(message.chat.id, 'Выложить в центр каждой лепешки по 1,5 ч. л. начинки, сложить кружок с начинкой пополам так, чтобы получился полумесяц. Соединить концы полумесяца и скрепить их. Плотно прижать пальцами, чтобы пельмени не разваливались при варке.')
			bot.send_sticker(message.chat.id,sh5)

			bot.send_message(message.chat.id, 'ШАГ 6')
			bot.send_message(message.chat.id, 'Выложить пельмени на поднос или плоское блюдо, присыпать мукой и поставить в холодильник. Так же приготовить пельмени из оставшегося теста.')
			bot.send_sticker(message.chat.id,sh6)

		elif message.text == 'Камень, ножницы, бумага':

			markup = types.InlineKeyboardMarkup(row_width=2)
			
			item1 = types.InlineKeyboardButton('Камень', callback_data='rock')
			item2 = types.InlineKeyboardButton('Ножницы', callback_data='scissors')
			item3 = types.InlineKeyboardButton('Бумага', callback_data='paper')
			markup.add(item1,item2,item3)

			journal = open('journal.txt', 'w')
			line_in_tabel = dictionary[str(user_name)]

			table = openpyxl.load_workbook('tables/'+str(user_name)+'.xlsx')
			
			sheet = table.active

			bot.send_message(channel_adress, '{} играет в "Камень, ножницы, бумага"'.format(user_name))

			sheet['A'+str(line_in_tabel)] = 'Играет в "Камень, ножницы, бумага"'
			sheet['B'+str(line_in_tabel)] = str(datetime.datetime.now())[0:19]

			table.save('tables/{}.xlsx'.format(user_name))

			line_in_tabel += 1
			dictionary[str(user_name)] = line_in_tabel
			
			journal.write(str(dictionary))
			journal.close()

			bot.send_message(message.chat.id, 'Твой выбор' , reply_markup=markup)

		elif message.text == 'Сообщение в канал':

			bot.send_message(message.chat.id, 'введите сообщение:')
			
			test_for_chanell = True
		elif test_for_chanell == True:
			bot.send_message(channel_adress, message.text)
			test_for_chanell = False
		else:
			bot.send_message(message.chat.id, 'Я не знаю что ответить 🤷')

@bot.callback_query_handler(func=lambda call: True)
def choose(call):
	
	rand = random.randint(0,1)
	rnd = random.randint(0,2)
	
	orel = open('st/orel.png','rb')
	reshka = open('st/reshka.png','rb')
	scissors = open ('st/scissors.png','rb')
	rock = open ('st/rock.png','rb')
	paper = open ('st/paper.png','rb')
	
	if call.message:
		if call.data == 'o':
			if rand == 0:
				bot.send_sticker(call.message.chat.id,orel)
				bot.send_message(call.message.chat.id, 'Ты выиграл(а)!👍')
			else:
				bot.send_sticker(call.message.chat.id,reshka)
				bot.send_message(call.message.chat.id, 'Ты проиграл(а)😢')
			bot.edit_message_text(chat_id=call.message.chat.id, message_id= call.message.message_id, text='Загадай сторону',
			reply_markup=None)
		elif call.data == 'r':
			if rand == 1:
				bot.send_sticker(call.message.chat.id,reshka)
				bot.send_message(call.message.chat.id, 'Ты выиграл(а)!👍')
			else:
				bot.send_sticker(call.message.chat.id,orel)
				bot.send_message(call.message.chat.id, 'Ты проиграл(а)😢')
			bot.edit_message_text(chat_id=call.message.chat.id, message_id= call.message.message_id, text='Загадай сторону',
			reply_markup=None)
		elif call.data == 'rock':
			if rnd == 0:
				bot.send_sticker(call.message.chat.id,rock)
				bot.send_message(call.message.chat.id, 'Ничья👌')
			elif rnd == 1:
				bot.send_sticker(call.message.chat.id,scissors)
				bot.send_message(call.message.chat.id, 'Ты выиграл(а)!👍')
			elif rnd == 2:
				bot.send_sticker(call.message.chat.id,paper)
				bot.send_message(call.message.chat.id, 'Ты проиграл(а)😢')
			bot.edit_message_text(chat_id=call.message.chat.id, message_id= call.message.message_id, text='Твой выбор',
			reply_markup=None)
		elif call.data == 'scissors':
			if rnd == 0:
				bot.send_sticker(call.message.chat.id,rock)
				bot.send_message(call.message.chat.id, 'Ты проиграл(а)😢')
			elif rnd == 1:
				bot.send_sticker(call.message.chat.id,scissors)
				bot.send_message(call.message.chat.id, 'Ничья👌')
			elif rnd == 2:
				bot.send_sticker(call.message.chat.id,paper)
				bot.send_message(call.message.chat.id, 'Ты выиграл(а)!👍')
			bot.edit_message_text(chat_id=call.message.chat.id, message_id= call.message.message_id, text='Твой выбор',
			reply_markup=None)
		elif call.data == 'paper':
			if rnd == 0:
				bot.send_sticker(call.message.chat.id,rock)
				bot.send_message(call.message.chat.id, 'Ты выиграл(а)!👍')
			elif rnd == 1:
				bot.send_sticker(call.message.chat.id,scissors)
				bot.send_message(call.message.chat.id, 'Ты проиграл(а)😢')
			elif rnd == 2:
				bot.send_sticker(call.message.chat.id,paper)
				bot.send_message(call.message.chat.id, 'Ничья👌')		
			bot.edit_message_text(chat_id=call.message.chat.id, message_id= call.message.message_id, text='Твой выбор',
			reply_markup=None)

#RUN
bot.polling (none_stop=True)